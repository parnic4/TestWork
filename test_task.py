import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import win32com.client as win32

CHROME_DRIVER_PATH = r'C:\TestWork\chromedriver-win64\chromedriver.exe'
URL = "https://www.moex.com/ru/derivatives/currency-rate.aspx?currency=USD_RUB"
DEFAULT_EMAIL_TO = "parnic4@gmail.com"  
EXCEL_FILE = "currency_data.xlsx"

START_PICKER_LABEL_XPATH = "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/form/div[2]/span/label"
END_PICKER_LABEL_XPATH   = "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/form/div[3]/span/label"
SHOW_BUTTON_XPATH        = "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/form/div[4]/button"
START_PICKER_DAY_PARENT  = "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/div[3]/div[4]"
END_PICKER_DAY_PARENT    = "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/div[3]/div[7]"

DEBUG_DIR = "debug"
os.makedirs(DEBUG_DIR, exist_ok=True)


def debug_save(driver, name):
    screenshot_path = os.path.join(DEBUG_DIR, f"{name}.png")
    html_path = os.path.join(DEBUG_DIR, f"{name}.html")
    driver.save_screenshot(screenshot_path)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    print(f"[DEBUG] Сохранены {screenshot_path} и {html_path}")


def click_month_option_xpath(driver, month_code_text="06 - Июнь"):
    calendar_month_dropdown = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, ".ui-dropdown.ui-calendar__dropdown.-opened")
        )
    )
    month_xpath = f'.//div[@class="ui-select-option__content" and normalize-space(text())="{month_code_text}"]'
    month_elem = WebDriverWait(calendar_month_dropdown, 10).until(
        lambda d: d.find_element(By.XPATH, month_xpath)
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", month_elem)
    month_elem.click()
    print(f"[DEBUG] Клик по месяцу: '{month_code_text}'")


def pick_month_strict_css(driver, label_xpath, month_code_text="06 - Июнь"):
    label = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, label_xpath))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", label)
    label.click()
    time.sleep(0.7)

    calendar_dropdown = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-dropdown.ui-calendar.-opened"))
    )

    try:
        arrow = calendar_dropdown.find_element(
            By.CSS_SELECTOR,
            '.ui-calendar__controls > .ui-group-item.ui-select.ui-calendar__select:first-child .ui-select__icon.-arrow'
        )
        arrow.click()
        time.sleep(0.7)
        print("[DEBUG] Клик по стрелке выбора месяца (CSS уникально)")
    except Exception:
        print("[DEBUG] Не найден уникальный .ui-select__icon.-arrow для месяца")
        debug_save(driver, "arrow_in_calendar_not_found")
        raise

    try:
        period_btn = calendar_dropdown.find_element(By.CSS_SELECTOR, ".ui-calendar__period")
        if period_btn.is_displayed():
            period_btn.click()
            time.sleep(0.7)
            print("[DEBUG] Клик по .ui-calendar__period для открытия месяцев")
    except Exception:
        pass

    click_month_option_xpath(driver, month_code_text)
    time.sleep(1)


def pick_date_by_xpaths(driver, label_xpath, day_parent_xpath, day_number, month_code_text="06 - Июнь"):
    pick_month_strict_css(driver, label_xpath, month_code_text)
    day_xpath = f"{day_parent_xpath}//div[contains(@class, 'ui-calendar__cell') and text()='{day_number}']"
    try:
        day_elem = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, day_xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", day_elem)
        day_elem.click()
        time.sleep(0.4)
    except Exception:
        debug_save(driver, f"day_{day_number}_not_found")
        print(f"[DEBUG] Не найден/не кликабелен день: {day_number} по xpath: {day_xpath}")
        raise


def fetch_currency_data(driver, start_label_xpath, end_label_xpath,
                        start_day_parent, end_day_parent,
                        start_day=1, end_day=30,
                        month_code_text="06 - Июнь"):
    pick_date_by_xpaths(
        driver, start_label_xpath, start_day_parent, start_day, month_code_text
    )
    pick_date_by_xpaths(
        driver, end_label_xpath, end_day_parent, end_day, month_code_text
    )

    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, SHOW_BUTTON_XPATH))).click()
    time.sleep(7)

    try:
        container = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, ".ui-container.-default")
            )
        )
        table = container.find_element(By.CSS_SELECTOR, ".ui-table__container table")
    except Exception:
        debug_save(driver, "table_not_found")
        print("[DEBUG] Не найдена таблица результатов")
        raise

    rows = table.find_elements(By.XPATH, './/tr[contains(@class,"ui-table-row") and contains(@class,"-interactive")]')
    data = []
    for row in rows:
        tds = row.find_elements(By.CSS_SELECTOR, "td.ui-table-cell")
        if len(tds) >= 5:
            date = tds[0].text.strip()
            value = tds[3].text.strip()
            time_val = tds[4].text.strip()
            data.append({
                'Дата': date,
                'Курс основного клиринга': value,
                'Время': time_val,
            })
    print(f"Собрано {len(data)} записей.")
    return data


def format_excel(file_name, row_count):
    wb = load_workbook(file_name)
    ws = wb.active

    for col_cells in ws.columns:
        column_letter = col_cells[0].column_letter
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0
        )
        ws.column_dimensions[column_letter].width = max_len + 2
    
    financial_rub_format = (
        '_-* #,##0.00\\ [$₽-419]_-;'     
        '_-* (#,##0.00)\\ [$₽-419]_-;'   
        '_-* "-"??\\ [$₽-419]_-;'        
        '_-@_- '                         
    )
    for col in ('B', 'E', 'G'):
        for r in range(2, row_count + 2):
            cell = ws[f'{col}{r}']
            if cell.value is not None:
                cell.number_format = financial_rub_format

    from numbers import Number
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, Number):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='justify', vertical='center')

    if row_count > 0:
        sum_row = row_count + 2
        ws.cell(row=sum_row, column=7).value = "Сумма:"
        
        ws.cell(row=sum_row, column=8).value = f"=SUM(G2:G{row_count + 1})"
        ws.cell(row=sum_row, column=8).number_format = financial_rub_format

    wb.save(file_name)


def send_email(file_name, row_count):
    row_count += 2
    """Отправляет письмо через локальный Outlook, запрашивая адрес получателя."""
 
    if row_count % 10 == 1 and row_count % 100 != 11:
        word = "строка"
    elif 2 <= row_count % 10 <= 4 and not (12 <= row_count % 100 <= 14):
        word = "строки"
    else:
        word = "строк"

    subject = f"Данные по валютам - {row_count} {word}"
    body = (
        f"В приложении файл с данными за период.\n\n"
        f"Количество строк в таблице: {row_count} {word}."
    )

    recipient = input(
        "Введите email получателя (оставьте пустым для отправки на адрес по умолчанию): "
    ).strip()
    if not recipient:
        recipient = DEFAULT_EMAIL_TO

    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)  
    mail.To = recipient
    mail.Subject = subject
    mail.Body = body
    abs_path = os.path.abspath(file_name)
    mail.Attachments.Add(Source=abs_path)

    mail.Display()

    send_choice = input("Отправить письмо? (y/n): ").strip().lower()
    if send_choice == 'y':
        mail.Send()
        print(f"Письмо отправлено на {recipient}")
    else:
        print("Отправка отменена пользователем.")


def main():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    service = Service(executable_path=CHROME_DRIVER_PATH)
    driver = webdriver.Chrome(service=service, options=options)
    try:
        driver.get(URL)
        try:
            WebDriverWait(driver, 8).until(EC.element_to_be_clickable(
                (By.XPATH, "/html/body/div[1]/div/div/div/div/div[1]/div/a[1]")
            )).click()
        except Exception:
            pass
        try:
            WebDriverWait(driver, 8).until(EC.element_to_be_clickable(
                (By.XPATH, "/html/body/div[8]/div[1]/div/div/div[2]/button")
            )).click()
        except Exception:
            pass
        print("Сбор данных для USD/RUB...")
        usd_data = fetch_currency_data(
            driver,
            START_PICKER_LABEL_XPATH, END_PICKER_LABEL_XPATH,
            START_PICKER_DAY_PARENT, END_PICKER_DAY_PARENT,
            start_day=1, end_day=30,
            month_code_text="06 - Июнь"
        )
        print("Смена валюты на JPY/RUB...")
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/form/div[1]/div[1]/div")
        )).click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[2]/div[6]/div/div/div[2]/div/div/div/div/div[5]/div[3]/div[1]/div[8]/div/div[1]")
        )).click()
        time.sleep(5)
        print("Сбор данных для JPY/RUB...")
        jpy_data = fetch_currency_data(
            driver,
            START_PICKER_LABEL_XPATH, END_PICKER_LABEL_XPATH,
            START_PICKER_DAY_PARENT, END_PICKER_DAY_PARENT,
            start_day=1, end_day=30,
            month_code_text="06 - Июнь"
        )
        min_len = min(len(usd_data), len(jpy_data))
        usd_data = usd_data[:min_len]
        jpy_data = jpy_data[:min_len]
        df = pd.DataFrame({
            'Дата USD/RUB': [d['Дата'] for d in usd_data],
            'Курс USD/RUB': [d['Курс основного клиринга'] for d in usd_data],
            'Время USD/RUB': [d['Время'] for d in usd_data],
            'Дата JPY/RUB': [d['Дата'] for d in jpy_data],
            'Курс JPY/RUB': [d['Курс основного клиринга'] for d in jpy_data],
            'Время JPY/RUB': [d['Время'] for d in jpy_data],
        })

        df['Курс USD/RUB'] = pd.to_numeric(df['Курс USD/RUB'], errors='coerce')
        df['Курс JPY/RUB'] = pd.to_numeric(df['Курс JPY/RUB'], errors='coerce')
        df['Результат'] = df['Курс USD/RUB'] / df['Курс JPY/RUB']

        df.to_excel(EXCEL_FILE, index=False)
        format_excel(EXCEL_FILE, len(df))
        send_email(EXCEL_FILE, len(df))
        print(f"Создан файл: {os.path.abspath(EXCEL_FILE)}")
        print("Процесс успешно завершён!")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
